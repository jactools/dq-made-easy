#!/usr/bin/env python3
"""Transform Logical Data Definitions.xlsx to OpenMetadata-ready CSV files.

This script normalizes a wide, relation-heavy LDD export into a compact set of
files that can be used for OpenMetadata import/automation workflows.
"""

from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Set, Tuple

import openpyxl


REQUIRED_HEADERS: Set[str] = {
    "Asset Id",
    "Full Name",
    "Name",
    "Domain",
    "Status",
    "Definition",
    "Logical Data Type",
    "Domain values",
    "REF Technical Name",
    "[Data Attribute] mapping to Physical Data Dictionary [Column] > Full Name",
}


@dataclass
class LddRow:
    asset_id: str
    full_name: str
    name: str
    domain: str
    status: str
    definition: str
    notes: str
    logical_data_type: str
    domain_values: str
    ref_technical_name: str
    technical_name: str
    data_set_attr_name: str
    physical_full_name: str
    physical_name: str
    leading_bde_name: str
    supporting_bde_name: str


def clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value).strip()


def split_physical_path(path: str) -> Tuple[str, str, str]:
    """Split `Schema > Table > Column` into parts.

    If fewer than 3 path segments exist, missing pieces are returned as empty.
    """
    if not path:
        return "", "", ""
    parts = [p.strip() for p in path.split(">") if p and p.strip()]
    if len(parts) >= 3:
        return parts[-3], parts[-2], parts[-1]
    if len(parts) == 2:
        return "", parts[0], parts[1]
    if len(parts) == 1:
        return "", "", parts[0]
    return "", "", ""


def get_header_map(headers: List[str]) -> Dict[str, int]:
    return {h: i for i, h in enumerate(headers)}


def get_cell(row: List[object], hmap: Dict[str, int], name: str) -> str:
    idx = hmap.get(name)
    if idx is None or idx >= len(row):
        return ""
    return clean(row[idx])


def iter_ldd_rows(xlsx_path: Path) -> Iterable[LddRow]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [clean(v) for v in first_row]
    hmap = get_header_map(headers)

    for values_tuple in ws.iter_rows(min_row=2, values_only=True):
        values = list(values_tuple)
        if not any(v not in (None, "") for v in values):
            continue

        yield LddRow(
            asset_id=get_cell(values, hmap, "Asset Id"),
            full_name=get_cell(values, hmap, "Full Name"),
            name=get_cell(values, hmap, "Name"),
            domain=get_cell(values, hmap, "Domain"),
            status=get_cell(values, hmap, "Status"),
            definition=get_cell(values, hmap, "Definition"),
            notes=get_cell(values, hmap, "Notes"),
            logical_data_type=get_cell(values, hmap, "Logical Data Type"),
            domain_values=get_cell(values, hmap, "Domain values"),
            ref_technical_name=get_cell(values, hmap, "REF Technical Name"),
            technical_name=get_cell(values, hmap, "Technical Name"),
            data_set_attr_name=get_cell(values, hmap, "[Data Attribute] defines [Data Set Attribute] > Name"),
            physical_full_name=get_cell(values, hmap, "[Data Attribute] mapping to Physical Data Dictionary [Column] > Full Name"),
            physical_name=get_cell(values, hmap, "[Data Attribute] mapping to Physical Data Dictionary [Column] > Name"),
            leading_bde_name=get_cell(values, hmap, "[Data Attribute] has a leading mapping to [Business Data Element] > Name"),
            supporting_bde_name=get_cell(values, hmap, "[Data Attribute] has a supporting mapping to [Business Data Element] > Name"),
        )


def workbook_matches_structure(xlsx_path: Path) -> bool:
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        ws = wb.worksheets[0]
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = {clean(v) for v in first_row if clean(v)}
        return REQUIRED_HEADERS.issubset(headers)
    except Exception:
        return False


def find_input_workbook(input_dir: Path) -> Path:
    candidates = [
        p
        for p in input_dir.glob("*.xlsx")
        if p.is_file() and not p.name.startswith("~$") and workbook_matches_structure(p)
    ]
    if not candidates:
        raise FileNotFoundError(
            f"No .xlsx workbook with expected LDD structure found in: {input_dir}"
        )
    return sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)[0]


def write_glossary_terms(rows: List[LddRow], out_csv: Path) -> int:
    seen: Dict[str, LddRow] = {}
    for row in rows:
        key = row.asset_id or row.full_name or row.name
        if not key:
            continue
        if key not in seen:
            seen[key] = row

    with out_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "asset_id",
                "glossary_name",
                "term_name",
                "display_name",
                "description",
                "logical_data_type",
                "status",
                "synonyms",
                "domain_values",
                "notes",
            ],
        )
        writer.writeheader()
        for row in seen.values():
            synonyms = sorted({s for s in [row.ref_technical_name, row.technical_name, row.data_set_attr_name] if s})
            writer.writerow(
                {
                    "asset_id": row.asset_id,
                    "glossary_name": row.domain or "Logical Data Dictionary",
                    "term_name": row.name,
                    "display_name": row.full_name or row.name,
                    "description": row.definition,
                    "logical_data_type": row.logical_data_type,
                    "status": row.status,
                    "synonyms": "|".join(synonyms),
                    "domain_values": row.domain_values,
                    "notes": row.notes,
                }
            )

    return len(seen)


def write_column_mappings(rows: List[LddRow], out_csv: Path, service_name: str, database_name: str) -> int:
    count = 0
    with out_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "asset_id",
                "logical_name",
                "logical_full_name",
                "schema_name",
                "table_name",
                "column_name",
                "openmetadata_column_fqn",
                "description",
                "logical_data_type",
                "physical_mapping_full_name",
            ],
        )
        writer.writeheader()
        for row in rows:
            if not row.physical_full_name:
                continue
            schema_name, table_name, column_name = split_physical_path(row.physical_full_name)
            if not column_name:
                continue
            fqn = ""
            if schema_name and table_name and column_name:
                fqn = f"{service_name}.{database_name}.{schema_name}.{table_name}.{column_name}"
            writer.writerow(
                {
                    "asset_id": row.asset_id,
                    "logical_name": row.name,
                    "logical_full_name": row.full_name,
                    "schema_name": schema_name,
                    "table_name": table_name,
                    "column_name": column_name,
                    "openmetadata_column_fqn": fqn,
                    "description": row.definition,
                    "logical_data_type": row.logical_data_type,
                    "physical_mapping_full_name": row.physical_full_name,
                }
            )
            count += 1
    return count


def write_bde_assignments(rows: List[LddRow], out_csv: Path) -> int:
    count = 0
    with out_csv.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "asset_id",
                "logical_name",
                "physical_mapping_full_name",
                "assignment_type",
                "business_data_element_name",
            ],
        )
        writer.writeheader()
        for row in rows:
            if row.leading_bde_name:
                writer.writerow(
                    {
                        "asset_id": row.asset_id,
                        "logical_name": row.name,
                        "physical_mapping_full_name": row.physical_full_name,
                        "assignment_type": "leading",
                        "business_data_element_name": row.leading_bde_name,
                    }
                )
                count += 1
            if row.supporting_bde_name:
                writer.writerow(
                    {
                        "asset_id": row.asset_id,
                        "logical_name": row.name,
                        "physical_mapping_full_name": row.physical_full_name,
                        "assignment_type": "supporting",
                        "business_data_element_name": row.supporting_bde_name,
                    }
                )
                count += 1
    return count


def write_summary(
    out_md: Path,
    source: Path,
    term_count: int,
    mapping_count: int,
    bde_count: int,
    service_name: str,
    database_name: str,
) -> None:
    out_md.write_text(
        "\n".join(
            [
                "# OpenMetadata LDD Transformation Summary",
                "",
                f"- Source workbook: `{source}`",
                f"- Glossary terms exported: `{term_count}`",
                f"- Physical column mappings exported: `{mapping_count}`",
                f"- BDE assignments exported: `{bde_count}`",
                "",
                "## Notes",
                "",
                "- `openmetadata_glossary_terms.csv`: normalized logical/business terms.",
                "- `openmetadata_column_mappings.csv`: parsed `Schema > Table > Column` path with generated OpenMetadata FQN.",
                "- `openmetadata_bde_assignments.csv`: leading/supporting business data element relationships.",
                "",
                "## OpenMetadata FQN Placeholder",
                "",
                f"Generated FQN pattern uses service/database placeholders: `{service_name}.{database_name}.<schema>.<table>.<column>`",
                "Update `--service-name` and `--database-name` values if needed.",
            ]
        ),
        encoding="utf-8",
    )


def main() -> None:
    repo_root = Path(__file__).resolve().parents[2]

    parser = argparse.ArgumentParser(description="Transform LDD workbook for OpenMetadata")
    parser.add_argument(
        "--input",
        default="",
        help="Optional path to source Excel workbook (if omitted, auto-detect from --input-dir)",
    )
    parser.add_argument(
        "--input-dir",
        default=str(repo_root / "dq-db" / "mock-data"),
        help="Directory used for auto-detecting source workbook",
    )
    parser.add_argument(
        "--output-dir",
        default=str(repo_root / "dq-db" / "mock-data" / "openmetadata-ready"),
        help="Directory for generated CSV/MD files",
    )
    parser.add_argument("--service-name", default="source_service", help="OpenMetadata service name placeholder")
    parser.add_argument("--database-name", default="source_database", help="OpenMetadata database name placeholder")
    args = parser.parse_args()

    source = Path(args.input).expanduser() if args.input else find_input_workbook(Path(args.input_dir).expanduser())
    if not source.exists():
        raise FileNotFoundError(f"Input workbook not found: {source}")

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    rows = list(iter_ldd_rows(source))

    term_count = write_glossary_terms(rows, out_dir / "openmetadata_glossary_terms.csv")
    mapping_count = write_column_mappings(
        rows,
        out_dir / "openmetadata_column_mappings.csv",
        service_name=args.service_name,
        database_name=args.database_name,
    )
    bde_count = write_bde_assignments(rows, out_dir / "openmetadata_bde_assignments.csv")
    write_summary(
        out_dir / "README.md",
        source=source,
        term_count=term_count,
        mapping_count=mapping_count,
        bde_count=bde_count,
        service_name=args.service_name,
        database_name=args.database_name,
    )

    print(f"Wrote: {out_dir / 'openmetadata_glossary_terms.csv'} ({term_count} rows)")
    print(f"Wrote: {out_dir / 'openmetadata_column_mappings.csv'} ({mapping_count} rows)")
    print(f"Wrote: {out_dir / 'openmetadata_bde_assignments.csv'} ({bde_count} rows)")
    print(f"Wrote: {out_dir / 'README.md'}")


if __name__ == "__main__":
    main()
